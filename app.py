# app.py — Render + Supabase deployment version

import os
import io
import logging
import qrcode
import csv
import zipfile
import textwrap
import tempfile
from io import BytesIO, StringIO
from datetime import datetime, timezone, timedelta

# East Africa Time (UTC+3) — used everywhere instead of now_eat()
EAT = timezone(timedelta(hours=3))
def now_eat() -> datetime:
    """Return current datetime in East Africa Time (UTC+3)."""
    return datetime.now(tz=EAT)
from functools import wraps
from urllib.parse import quote as url_encode
from whatsapp import send_guest_card
from sms_africastalking import send_sms as at_send_sms, is_configured as at_configured
import time

from flask import (
    Flask, render_template, request, redirect, url_for, flash,
    session, jsonify, send_file, make_response, current_app
)

from werkzeug.utils import secure_filename
from dotenv import dotenv_values, load_dotenv
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from sqlalchemy.sql import func
from sqlalchemy.exc import IntegrityError

from supabase import create_client, Client
from models import Guest, Event, init_db, get_db_session
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether
    )
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ---------------------------------------------------------------------------
# Environment Loading
# ---------------------------------------------------------------------------

flask_env = os.getenv('FLASK_ENV', 'production')

WHATSAPP_PHONE_NUMBER_ID = os.getenv("WHATSAPP_PHONE_NUMBER_ID")
WHATSAPP_ACCESS_TOKEN    = os.getenv("WHATSAPP_ACCESS_TOKEN")
WHATSAPP_VERIFY_TOKEN    = os.getenv("WHATSAPP_VERIFY_TOKEN", "wedding_webhook_secret")

if flask_env == 'development':
    current_env_file = '.env.development'
else:
    current_env_file = '.env'

if os.path.exists(current_env_file):
    config = dotenv_values(current_env_file)
    for key, value in config.items():
        if value is not None:
            os.environ[key] = value
else:
    logging.warning(f"Env file {current_env_file} not found; using existing env vars.")

# ---------------------------------------------------------------------------
# Database Configuration
# ---------------------------------------------------------------------------

DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    DB_FILE      = os.getenv("DB_FILE", "guests.db")
    DATABASE_URL = f"sqlite:///./{DB_FILE}"

if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

# ---------------------------------------------------------------------------
# Supabase Storage Configuration
# ---------------------------------------------------------------------------

SUPABASE_URL  = os.getenv("SUPABASE_URL")
SUPABASE_KEY  = os.getenv("SUPABASE_SERVICE_KEY")
QR_BUCKET     = os.getenv("SUPABASE_QR_BUCKET",    "qr-codes")
CARDS_BUCKET     = os.getenv("SUPABASE_CARDS_BUCKET",    "guest-cards")
TEMPLATES_BUCKET = os.getenv("SUPABASE_TEMPLATES_BUCKET","card-templates")

# ---------------------------------------------------------------------------
# Default Event seed (used only when first event is auto-created from .env)
# ---------------------------------------------------------------------------
DEFAULT_WEDS_NAMES = os.getenv("EVENT_WEDS_NAMES", "the Bride & Groom")
DEFAULT_EVENT_DAY  = os.getenv("EVENT_DAY",        "Jumamosi")
DEFAULT_EVENT_DATE = os.getenv("EVENT_DATE",       "TBD")
DEFAULT_EVENT_VENUE= os.getenv("EVENT_VENUE",      "TBD")

supabase: Client = None
if SUPABASE_URL and SUPABASE_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    logging.info("Supabase client initialized.")
else:
    logging.warning("SUPABASE_URL or SUPABASE_SERVICE_KEY not set — storage features disabled.")

# ---------------------------------------------------------------------------
# Flask App
# ---------------------------------------------------------------------------

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')
if not app.config['SECRET_KEY']:
    raise ValueError("SECRET_KEY environment variable is not set.")

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL

UPLOAD_FOLDER = "uploads"
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logging.info(f"Using database: {DATABASE_URL}")

# Credentials — set ALL of these in Render environment variables
# No hardcoded defaults for security — app will reject login if not set
ADMIN_USERNAME  = os.environ.get("ADMIN_USERNAME")
ADMIN_PASSWORD  = os.environ.get("ADMIN_PASSWORD")
WORKER_USERNAME = os.environ.get("WORKER_USERNAME")
WORKER_PASSWORD = os.environ.get("WORKER_PASSWORD")

if not ADMIN_USERNAME or not ADMIN_PASSWORD:
    logging.warning("ADMIN_USERNAME or ADMIN_PASSWORD not set in environment variables!")

with app.app_context():
    init_db(app)


@app.context_processor
def inject_events():
    """Make active_event_name and all_events available in every template."""
    if not session.get('logged_in'):
        return {}
    try:
        with get_db_session() as db:
            ev  = get_active_event(db)
            evs = db.query(Event).filter_by(is_active=True).order_by(Event.id).all()
            all_evs = [{'id': e.id, 'name': e.name} for e in evs]
            return {
                'active_event_name': ev.name if ev else 'No Event',
                'all_events':        all_evs,
            }
    except Exception:
        return {}

# ---------------------------------------------------------------------------
# Card rendering constants  (1240 × 1748 px template)
# ---------------------------------------------------------------------------

_MONTSERRAT_PATH = os.path.join("static", "fonts", "Montserrat-Bold.ttf")
_POPPINS_PATH    = "/usr/share/fonts/truetype/google-fonts/Poppins-Bold.ttf"
_ROBOTO_PATH     = os.path.join("static", "fonts", "Roboto-Bold.ttf")

def _bold_font_path() -> str:
    for p in (_MONTSERRAT_PATH, _ROBOTO_PATH, _POPPINS_PATH):
        if os.path.exists(p):
            return p
    raise FileNotFoundError("No bold font found. Place Montserrat-Bold.ttf in static/fonts/.")

# Name — dotted line (left half of template)
NAME_CENTER_X  = 303
NAME_DOTTED_Y  = 550   # tweak if name sits above/below the line
NAME_MAX_WIDTH = 358

# QR code — bottom-left
QR_SIZE   = 200
QR_MARGIN = 45
QR_X      = QR_MARGIN
QR_Y      = 1748 - QR_SIZE - QR_MARGIN   # 1503

# Card number — top-left corner
CARD_NUM_COLOR = "#185a3f"
CARD_NUM_SIZE  = 38
CARD_NUM_TOP_X = 45
CARD_NUM_TOP_Y = 50

# Card type — above the QR code
CARD_TYPE_COLOR = "#185a3f"
CARD_TYPE_SIZE  = 36
CARD_TYPE_GAP   = 20

# ---------------------------------------------------------------------------
# Supabase helpers
# ---------------------------------------------------------------------------

def upload_to_supabase(bucket: str, filename: str, data: bytes,
                       content_type: str = "image/jpeg") -> str:
    if not supabase:
        raise RuntimeError("Supabase client not initialized.")
    supabase.storage.from_(bucket).upload(
        path=filename,
        file=data,
        file_options={"content-type": content_type, "upsert": "true"},
    )
    return supabase.storage.from_(bucket).get_public_url(filename)

def delete_from_supabase(bucket: str, filename: str):
    if not supabase:
        return
    try:
        supabase.storage.from_(bucket).remove([filename])
    except Exception as e:
        logging.warning(f"Could not delete {filename} from {bucket}: {e}")

def download_from_supabase(bucket: str, filename: str) -> bytes:
    if not supabase:
        raise RuntimeError("Supabase client not initialized.")
    return supabase.storage.from_(bucket).download(filename)

def qr_filename_from_guest(guest) -> str:
    sanitized = get_safe_filename_name_part(guest.name or "GUEST")
    return f"{guest.qr_code_id}-{sanitized}.png"

def card_filename_from_guest(guest) -> str:
    sanitized = get_safe_filename_name_part(guest.name or "GUEST")
    return f"GUEST-{guest.visual_id:04d}-{sanitized}.jpg"

# ---------------------------------------------------------------------------
# QR Code Generation
# ---------------------------------------------------------------------------

def generate_qr_bytes(data: str) -> bytes:
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf.getvalue()

# ---------------------------------------------------------------------------
# Card Rendering
# ---------------------------------------------------------------------------

def _fit_name_font(draw, name: str, max_width: int):
    font_path = _bold_font_path()
    for size in range(44, 19, -2):
        font  = ImageFont.truetype(font_path, size)
        bbox  = font.getbbox(name)
        if bbox[2] - bbox[0] <= max_width:
            return font, [name]
        wrapped = textwrap.fill(name, width=20)
        lines   = wrapped.split("\n")
        widths  = [(font.getbbox(l)[2] - font.getbbox(l)[0]) for l in lines]
        if max(widths) <= max_width:
            return font, lines
    font    = ImageFont.truetype(font_path, 22)
    wrapped = textwrap.fill(name, width=22)
    return font, wrapped.split("\n")


def _draw_card(guest, qr_img: Image.Image, template_bytes=None) -> Image.Image:
    """
    Card layout:
    1. Card number  → top-left  (NO. XXXX)
    2. Guest name   → centred on dotted-line placeholder
    3. QR code      → bottom-left
    4. Card type    → above QR  (SINGLE / DOUBLE / FAMILY)
    """
    # Per-event template takes priority over the global static one
    if template_bytes:
        img = Image.open(BytesIO(template_bytes)).convert("RGB")
    else:
        template_path = os.path.join("static", "Card Template.jpg")
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Card template not found: {template_path}")
        img = Image.open(template_path).convert("RGB")
    draw = ImageDraw.Draw(img)
    fp   = _bold_font_path()

    num_font  = ImageFont.truetype(fp, CARD_NUM_SIZE)
    type_font = ImageFont.truetype(fp, CARD_TYPE_SIZE)

    # 1. Card number — top-left
    draw.text((CARD_NUM_TOP_X, CARD_NUM_TOP_Y),
              f"NO. {guest.visual_id:04d}",
              font=num_font, fill=CARD_NUM_COLOR)

    # 2. Guest name — dotted line
    raw_name    = (guest.name or "GUEST").upper()
    name_font, lines = _fit_name_font(draw, raw_name, NAME_MAX_WIDTH)
    sample_bbox = name_font.getbbox("Ag")
    font_h      = sample_bbox[3] - sample_bbox[1]
    line_h      = font_h + 6
    total_h     = line_h * len(lines) - 6
    block_top_y = NAME_DOTTED_Y - total_h // 2
    for i, line in enumerate(lines):
        bbox   = draw.textbbox((0, 0), line, font=name_font)
        text_w = bbox[2] - bbox[0]
        draw.text((NAME_CENTER_X - text_w // 2, block_top_y + i * line_h),
                  line, font=name_font, fill="#000000")

    # 3. QR code — bottom-left
    img.paste(qr_img.resize((QR_SIZE, QR_SIZE), Image.LANCZOS), (QR_X, QR_Y))

    # 4. Card type — above QR
    type_label = "GROUP" if (guest.card_type or "").lower() == "family" else (guest.card_type or "SINGLE").upper()
    type_bbox  = draw.textbbox((0, 0), type_label, font=type_font)
    draw.text((QR_X, QR_Y - (type_bbox[3] - type_bbox[1]) - CARD_TYPE_GAP),
              type_label, font=type_font, fill=CARD_TYPE_COLOR)

    return img


def _render_and_upload_card(guest) -> bool:
    """Render JPEG card for one guest and upload to Supabase. Returns True on success."""
    try:
        # Try fetching the QR; fall back to regenerating if name was edited
        try:
            qr_data = download_from_supabase(QR_BUCKET, qr_filename_from_guest(guest))
        except Exception:
            qr_data = generate_qr_bytes(guest.qr_code_id)

        qr_img = Image.open(BytesIO(qr_data))
        img    = _draw_card(guest, qr_img)
        buf    = BytesIO()
        img.save(buf, format="JPEG", quality=92)
        buf.seek(0)
        upload_to_supabase(CARDS_BUCKET, card_filename_from_guest(guest),
                           buf.getvalue(), content_type="image/jpeg")
        return True
    except Exception as e:
        logging.error(f"_render_and_upload_card failed for {guest.name}: {e}")
        return False


def _get_event_template_bytes(event) -> bytes | None:
    """Fetch per-event card template from Supabase, or None to use global static."""
    if not event or not event.card_template_url:
        return None
    try:
        fname = f"template_{event.slug}.jpg"
        return download_from_supabase(TEMPLATES_BUCKET, fname)
    except Exception:
        return None


def _generate_card_bytes(guest, event=None) -> bytes | None:
    """Return raw JPEG card bytes (used as fallback in send engine)."""
    try:
        try:
            qr_data = download_from_supabase(QR_BUCKET, qr_filename_from_guest(guest))
        except Exception:
            qr_data = generate_qr_bytes(guest.qr_code_id)

        template_bytes = _get_event_template_bytes(event)
        qr_img = Image.open(BytesIO(qr_data))
        img    = _draw_card(guest, qr_img, template_bytes=template_bytes)
        buf    = BytesIO()
        img.save(buf, format="JPEG", quality=92)
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        logging.error(f"_generate_card_bytes failed for {guest.name}: {e}")
        return None

# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def to_whatsapp_number(phone):
    phone = str(phone).strip()
    if phone.startswith('+'): phone = phone[1:]
    if phone.startswith('255') and len(phone) == 12: return phone
    if phone.startswith('0'): phone = phone[1:]
    if len(phone) == 9 and phone[0] in ('7', '6'): return f"255{phone}"
    if phone.startswith('255'): return phone
    return phone

app.jinja_env.globals.update(to_whatsapp_number=to_whatsapp_number, url_encode=url_encode)

def get_safe_filename_name_part(name):
    return "".join(c if c.isalnum() else '_' for c in (name or "").upper())

def normalize_card_type(card_type_input, allowed_input=None):
    card_type = (card_type_input or "").strip().lower()
    if card_type in ("s", "single"):        return "single", 1
    if card_type in ("d", "double"):        return "double", 2
    if card_type in ("f", "family", "group"):
        if allowed_input:
            try:
                a = int(allowed_input)
                if a >= 3: return "family", a
                if a == 2: return "double", 2
            except ValueError: pass
        return "family", 5
    if allowed_input:
        try:
            a = int(allowed_input)
            if a <= 1: return "single", 1
            if a == 2: return "double", 2
            return "family", a
        except ValueError: pass
    return "single", 1

def get_next_visual_id(db_session, event_id=None):
    """Return next unique visual_id.
    Per-event when event_id given; falls back to global max to avoid DB collisions.
    We offset per-event IDs by event_id * 10000 to guarantee global uniqueness
    while still having per-event numbering that starts near 1.
    E.g. Event 1 → 10001, 10002 ... Event 2 → 20001, 20002 ...
    Exception: event_id=1 (default/first event) keeps original numbering (1,2,3...)
    for backwards compatibility with existing guests."""
    if event_id is None:
        max_id = db_session.query(func.max(Guest.visual_id)).scalar()
        return 1 if max_id is None else int(max_id) + 1
    if event_id == 1:
        # Legacy first event — keep 1-based numbering
        q = db_session.query(func.max(Guest.visual_id)).filter(
            Guest.event_id == event_id)
        max_id = q.scalar()
        return 1 if max_id is None else int(max_id) + 1
    # Other events: base offset = event_id * 10000
    base   = event_id * 10000
    q      = db_session.query(func.max(Guest.visual_id)).filter(
        Guest.visual_id >= base,
        Guest.visual_id <  base + 10000,
    )
    max_id = q.scalar()
    return base + 1 if max_id is None else int(max_id) + 1

# Swahili event type labels for SMS
EVENT_TYPE_LABELS = {
    "Wedding":      "HARUSI",
    "Send-Off":     "SEND-OFF",
    "Birthday":     "SIKUKUU YA KUZALIWA",
    "Conference":   "MKUTANO",
    "Confirmation": "IBADA YA KIPAIMARA",
    "Corporate":    "TUKIO LA KAMPUNI",
    "Other":        "TUKIO",
}

def build_sms_message(guest, event=None) -> str:
    weds       = (event.weds_names  if event and event.weds_names  else DEFAULT_WEDS_NAMES) or ""
    day        = (event.event_day   if event and event.event_day   else DEFAULT_EVENT_DAY)  or ""
    date       = (event.event_date  if event and event.event_date  else DEFAULT_EVENT_DATE) or ""
    venue      = (event.event_venue if event and event.event_venue else DEFAULT_EVENT_VENUE)or ""
    ev_type    = (event.event_type  if event and event.event_type  else "Wedding")
    type_label = EVENT_TYPE_LABELS.get(ev_type, ev_type.upper())
    return (
        f"MWALIKO\n"
        f"Habari {guest.name},\n"
        f"Umealikwa {type_label} ya:\n"
        f"{weds.upper()}\n"
        f"{day.upper()}, {date.upper()}\n"
        f"Saa 12:00 Jioni\n"
        f"{venue.upper()}\n"
        f"\n"
        f"Kadi No: {guest.visual_id:04d} - {(guest.card_type or 'Single').title()}\n"
        f"Fika na kadi hii ukumbini.\n"
        f"Karibu sana! - SwiftInvite"
    )


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            flash('Please log in first.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """Restrict route to admin role only."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            flash('Please log in first.', 'warning')
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('view_all'))
        return f(*args, **kwargs)
    return decorated_function

# ---------------------------------------------------------------------------
# Event helpers — active event in session
# ---------------------------------------------------------------------------

def _seed_default_event(db):
    """Create the default event from .env values if no events exist yet."""
    import re
    if db.query(Event).count() == 0:
        slug = re.sub(r'[^a-z0-9]+', '-',
                      DEFAULT_WEDS_NAMES.lower()).strip('-') or 'event-1'
        ev = Event(
            name          = DEFAULT_WEDS_NAMES,
            slug          = slug,
            weds_names    = DEFAULT_WEDS_NAMES,
            event_day     = DEFAULT_EVENT_DAY,
            event_date    = DEFAULT_EVENT_DATE,
            event_venue   = DEFAULT_EVENT_VENUE,
            storage_prefix= slug,
            is_active     = True,
            created_at    = now_eat(),
        )
        db.add(ev)
        db.flush()
        # Assign all existing guests to this event
        db.query(Guest).filter(Guest.event_id.is_(None)).update(
            {'event_id': ev.id}, synchronize_session=False
        )
        db.commit()
        return ev
    return None


def get_active_event(db):
    """Return the Event the user is currently working on.
    Falls back to the first event if nothing is set in session."""
    _seed_default_event(db)
    eid = session.get('active_event_id')
    ev  = None
    if eid:
        ev = db.get(Event, eid)
    if not ev:
        ev = db.query(Event).order_by(Event.id).first()
        if ev:
            session['active_event_id'] = ev.id
    return ev


def require_event(f):
    """Decorator: ensure an active event is in session before entering route."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        with get_db_session() as db:
            ev = get_active_event(db)
            if not ev:
                flash('Please create an event first.', 'warning')
                return redirect(url_for('events_list'))
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Event management routes
# ---------------------------------------------------------------------------

@app.route('/events')
@login_required
def events_list():
    with get_db_session() as db:
        _seed_default_event(db)
        events = db.query(Event).order_by(Event.id.desc()).all()
        active = get_active_event(db)
        # snapshot so objects are usable outside session
        evs = [{
            'id':           e.id,
            'name':         e.name,
            'slug':         e.slug,
            'weds_names':   e.weds_names,
            'event_day':    e.event_day,
            'event_date':   e.event_date,
            'event_venue':  e.event_venue,
            'is_active':    e.is_active,
            'guest_count':  db.query(Guest).filter_by(event_id=e.id).count(),
        } for e in events]
        active_id = active.id if active else None
    return render_template('events.html', events=evs, active_id=active_id)


@app.route('/events/new', methods=['GET', 'POST'])
@admin_required
def event_new():
    if request.method == 'POST':
        import re
        name   = request.form.get('name', '').strip()
        if not name:
            flash('Event name is required.', 'danger')
            return redirect(url_for('event_new'))
        slug_base = re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')
        with get_db_session() as db:
            # Ensure slug is unique
            slug = slug_base
            counter = 1
            while db.query(Event).filter_by(slug=slug).first():
                slug = f"{slug_base}-{counter}"
                counter += 1
            ev = Event(
                name           = name,
                slug           = slug,
                event_type     = request.form.get('event_type', 'Wedding').strip(),
                weds_names     = request.form.get('weds_names', '').strip(),
                event_day      = request.form.get('event_day',  '').strip(),
                event_date     = request.form.get('event_date', '').strip(),
                event_venue    = request.form.get('event_venue','').strip(),
                wa_phone_number_id   = request.form.get('wa_phone_number_id','').strip() or None,
                wa_access_token      = request.form.get('wa_access_token','').strip() or None,
                wa_template_name     = request.form.get('wa_template_name','event_invitation').strip() or 'event_invitation',
                wa_template_language = request.form.get('wa_template_language','sw').strip() or 'sw',
                at_username    = request.form.get('at_username','').strip() or None,
                at_api_key     = request.form.get('at_api_key','').strip() or None,
                at_sender_id   = request.form.get('at_sender_id','').strip() or None,
                storage_prefix = slug,
                is_active      = True,
                created_at     = now_eat(),
            )
            db.add(ev)
            db.commit()
            db.refresh(ev)
            session['active_event_id'] = ev.id
            flash(f'Event "{name}" created and set as active.', 'success')
        return redirect(url_for('events_list'))
    return render_template('event_form.html', event=None, title='New Event')


@app.route('/events/<int:event_id>/edit', methods=['GET', 'POST'])
@admin_required
def event_edit(event_id):
    with get_db_session() as db:
        ev = db.get(Event, event_id)
        if not ev:
            flash('Event not found.', 'danger')
            return redirect(url_for('events_list'))
        if request.method == 'POST':
            ev.name           = request.form.get('name', ev.name).strip()
            ev.event_type     = request.form.get('event_type', ev.event_type or 'Wedding').strip()
            ev.weds_names     = request.form.get('weds_names', '').strip()
            ev.event_day      = request.form.get('event_day',  '').strip()
            ev.event_date     = request.form.get('event_date', '').strip()
            ev.event_venue    = request.form.get('event_venue','').strip()
            ev.wa_phone_number_id   = request.form.get('wa_phone_number_id','').strip() or None
            ev.wa_access_token      = request.form.get('wa_access_token','').strip() or None
            ev.wa_template_name     = request.form.get('wa_template_name','event_invitation').strip() or 'event_invitation'
            ev.wa_template_language = request.form.get('wa_template_language','sw').strip() or 'sw'
            ev.at_username    = request.form.get('at_username','').strip() or None
            ev.at_api_key     = request.form.get('at_api_key','').strip() or None
            ev.at_sender_id   = request.form.get('at_sender_id','').strip() or None
            db.commit()
            flash(f'Event "{ev.name}" updated.', 'success')
            return redirect(url_for('events_list'))
        # Snapshot for template
        ev_data = {
            'id': ev.id, 'name': ev.name, 'slug': ev.slug,
            'event_type': ev.event_type or 'Wedding',
            'card_template_url': ev.card_template_url or '',
            'weds_names': ev.weds_names, 'event_day': ev.event_day,
            'event_date': ev.event_date, 'event_venue': ev.event_venue,
            'wa_phone_number_id': ev.wa_phone_number_id or '',
            'wa_access_token': ev.wa_access_token or '',
            'wa_template_name': ev.wa_template_name or 'event_invitation',
            'wa_template_language': ev.wa_template_language or 'sw',
            'at_username': ev.at_username or '',
            'at_api_key': ev.at_api_key or '',
            'at_sender_id': ev.at_sender_id or '',
        }
    return render_template('event_form.html', event=ev_data, title='Edit Event')


@app.route('/events/<int:event_id>/switch')
@login_required
def event_switch(event_id):
    with get_db_session() as db:
        ev = db.get(Event, event_id)
        if not ev:
            flash('Event not found.', 'danger')
        else:
            session['active_event_id'] = ev.id
            flash(f'Switched to: {ev.name}', 'success')
    return redirect(request.referrer or url_for('view_all'))


@app.route('/events/<int:event_id>/archive', methods=['POST'])
@admin_required
def event_archive(event_id):
    with get_db_session() as db:
        ev = db.get(Event, event_id)
        if ev:
            ev.is_active = not ev.is_active
            db.commit()
            state = 'reactivated' if ev.is_active else 'archived'
            flash(f'Event "{ev.name}" {state}.', 'success')
    return redirect(url_for('events_list'))


@app.route('/events/<int:event_id>/upload_template', methods=['POST'])
@admin_required
def event_upload_template(event_id):
    """Upload a per-event card template image to Supabase."""
    file = request.files.get('template_file')
    if not file or not file.filename:
        flash('No file selected.', 'danger')
        return redirect(url_for('event_edit', event_id=event_id))
    if not file.filename.lower().endswith(('.jpg', '.jpeg', '.png')):
        flash('Only JPG/PNG files are supported.', 'danger')
        return redirect(url_for('event_edit', event_id=event_id))
    with get_db_session() as db:
        ev = db.get(Event, event_id)
        if not ev:
            flash('Event not found.', 'danger')
            return redirect(url_for('events_list'))
        fname = f"template_{ev.slug}.jpg"
        img_bytes = file.read()
        # Convert PNG to JPEG if needed
        if file.filename.lower().endswith('.png'):
            from PIL import Image as PilImage
            img_obj = PilImage.open(BytesIO(img_bytes)).convert('RGB')
            buf = BytesIO()
            img_obj.save(buf, format='JPEG', quality=95)
            img_bytes = buf.getvalue()
        url = upload_to_supabase(TEMPLATES_BUCKET, fname, img_bytes,
                                  content_type='image/jpeg')
        ev.card_template_url = url
        db.commit()
        flash(f'Card template uploaded for "{ev.name}".', 'success')
    return redirect(url_for('event_edit', event_id=event_id))



# ---------------------------------------------------------------------------
# Routes — core
# ---------------------------------------------------------------------------

@app.route('/')
@login_required
def view_all():
    with get_db_session() as db:
        ev      = get_active_event(db)
        eid     = ev.id if ev else None
        guests  = db.query(Guest).filter_by(event_id=eid).order_by(Guest.visual_id).all()
        missing = [g for g in guests if g.visual_id is None]
        for g in missing:
            g.visual_id = get_next_visual_id(db)
            db.add(g)
        if missing:
            db.commit()
        ev_name = ev.name if ev else "No Event"
    return render_template('guests.html', guests=guests,
                           current_environment=flask_env, active_event_name=ev_name)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        # Read fresh from env each login attempt
        _admin_u = os.environ.get("ADMIN_USERNAME")
        _admin_p = os.environ.get("ADMIN_PASSWORD")
        _worker_u= os.environ.get("WORKER_USERNAME")
        _worker_p= os.environ.get("WORKER_PASSWORD")
        if _admin_u and _admin_p and username == _admin_u and password == _admin_p:
            session['logged_in'] = True
            session['role']      = 'admin'
            flash('Login successful.', 'success')
            return redirect(url_for('view_all'))
        elif (_worker_u and _worker_p
              and username == _worker_u
              and password == _worker_p):
            session['logged_in'] = True
            session['role']      = 'worker'
            flash('Logged in as worker.', 'success')
            return redirect(url_for('view_all'))
        flash('Invalid credentials.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    session.pop('logged_in', None)
    session.pop('role', None)
    flash('Logged out.', 'info')
    return redirect(url_for('login'))

# -------------------- add_guest --------------------

@app.route('/add_guest', methods=['GET', 'POST'])
@admin_required
def add_guest():
    if request.method == 'POST':
        name             = (request.form.get('name') or '').strip()
        phone            = to_whatsapp_number((request.form.get('phone') or '').strip())
        card_type_input  = request.form.get('card_type', 'single')
        group_size_input = request.form.get('group_size', '').strip()
        card_type, default_size = normalize_card_type(card_type_input, group_size_input)
        group_size = (int(group_size_input)
                      if card_type == 'family' and group_size_input.isdigit()
                      else default_size)

        with get_db_session() as db:
            ev        = get_active_event(db)
            eid       = ev.id if ev else None
            if db.query(Guest).filter_by(phone=phone, event_id=eid).first():
                flash(f"Guest with phone {phone} already exists in this event.", "warning")
                return redirect(url_for('add_guest'))
            visual_id = get_next_visual_id(db, eid)
            qr_id     = f"EV{eid or 0}-GUEST-{visual_id:04d}"
            try:
                qr_bytes = generate_qr_bytes(qr_id)
                qr_fname = f"{qr_id}-{get_safe_filename_name_part(name or 'GUEST')}.png"
                qr_url   = upload_to_supabase(QR_BUCKET, qr_fname, qr_bytes,
                                              content_type="image/png")
            except Exception as e:
                current_app.logger.warning(f"QR upload failed: {e}")
                qr_url = ""

            ev = get_active_event(db)
            db.add(Guest(name=name, phone=phone, qr_code_id=qr_id,
                         qr_code_url=qr_url, visual_id=visual_id,
                         card_type=card_type, group_size=group_size,
                         checked_in_count=0,
                         event_id=ev.id if ev else None))
            db.commit()
            flash(f"Guest '{name or phone}' added. Card: {card_type.title()}, "
                  f"entries: {group_size}.", "success")
            return redirect(url_for('view_all'))

    return render_template('add_guest.html')

# -------------------- upload_csv --------------------

@app.route('/upload_csv', methods=['GET', 'POST'])
@admin_required
def upload_csv():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash("No file selected.", "danger")
            return redirect(request.url)

        stream = StringIO(file.stream.read().decode("utf-8"))
        reader = csv.DictReader(stream)
        added = skipped = 0

        def get_row(row, *keys):
            for k in keys:
                v = row.get(k) or row.get(k.lower()) or row.get(k.capitalize())
                if v: return v.strip()
            return ""

        def normalize(raw):
            raw = (raw or "").strip().lower()
            if raw in ["s", "single"]:          return "single"
            if raw in ["d", "double"]:          return "double"
            if raw in ["f", "family", "group"]: return "family"
            return "single"

        with get_db_session() as db:
            ev  = get_active_event(db)
            eid = ev.id if ev else None
            for row in reader:
                name      = get_row(row, "name", "Name")
                raw_phone = get_row(row, "phone", "Phone")
                if not raw_phone:
                    skipped += 1; continue

                phone     = to_whatsapp_number(raw_phone)
                card_type = normalize(get_row(row, "Card Type", "card_type", "type"))
                if card_type == "single":   group_size = 1
                elif card_type == "double": group_size = 2
                else:
                    try:
                        group_size = max(1, int(get_row(
                            row, "Allowed", "allowed", "Size", "size",
                            "Group Size", "group_size")))
                    except Exception:
                        group_size = 1

                if db.query(Guest).filter_by(phone=phone, event_id=eid).first():
                    skipped += 1; continue

                visual_id = get_next_visual_id(db, eid)
                qr_id     = f"EV{eid or 0}-GUEST-{visual_id:04d}"
                qr_fname  = f"{qr_id}-{get_safe_filename_name_part(name or 'GUEST')}.png"
                try:
                    qr_bytes = generate_qr_bytes(qr_id)
                    qr_url   = upload_to_supabase(QR_BUCKET, qr_fname, qr_bytes,
                                                  content_type="image/png")
                except Exception as e:
                    current_app.logger.warning(f"QR upload failed for {name}: {e}")
                    qr_url = ""

                db.add(Guest(name=name, phone=phone, qr_code_id=qr_id,
                             qr_code_url=qr_url, visual_id=visual_id,
                             card_type=card_type, group_size=group_size,
                             checked_in_count=0, event_id=eid))
                db.flush()
                added += 1
            db.commit()

        flash(f"CSV processed — Added: {added}, Skipped: {skipped}", "success")
        return redirect(url_for('view_all'))

    return render_template('upload_csv.html')

# -------------------- update_status --------------------

@app.route('/update_status', methods=['POST'])
@login_required
def update_status():
    data       = request.get_json() or {}
    qr_code_id = data.get("qr_code_id")
    visual_id  = data.get("visual_id")
    if not qr_code_id and not visual_id:
        return jsonify(success=False, message="Missing qr_code_id or visual_id.")

    with get_db_session() as db:
        try:
            ev    = get_active_event(db)
            eid   = ev.id if ev else None
            guest = None
            if qr_code_id:
                guest = db.query(Guest).filter_by(qr_code_id=qr_code_id).first()
            if not guest and visual_id:
                guest = db.query(Guest).filter_by(
                    visual_id=int(visual_id), event_id=eid).first()
            if not guest:
                return jsonify(success=False, message="Guest not found.")
            remaining = guest.group_size - guest.checked_in_count
            if remaining <= 0:
                return jsonify(
                    success=False, already_entered=True,
                    message="All allowed entries have already checked in.",
                    guest={"visual_id": guest.visual_id, "name": guest.name,
                           "card_type": (guest.card_type or "").title(),
                           "remaining_entries": 0})
            guest.checked_in_count = (guest.checked_in_count or 0) + 1
            if guest.checked_in_count >= guest.group_size:
                guest.has_entered = True
                guest.entry_time  = now_eat()
            db.commit()
            return jsonify(
                success=True, message="Check-in successful.",
                guest={"visual_id": guest.visual_id, "name": guest.name,
                       "card_type": (guest.card_type or "").title(),
                       "remaining_entries": guest.group_size - guest.checked_in_count})
        except Exception as e:
            db.rollback()
            current_app.logger.exception(f"Error updating status for {qr_code_id}: {e}")
            return jsonify(success=False, message=f"An error occurred: {e}")

# -------------------- search_guests --------------------

@app.route('/search_guests')
@login_required
def search_guests():
    query = request.args.get('q', '').strip()
    with get_db_session() as db:
        ev  = get_active_event(db)
        eid = ev.id if ev else None
        if query:
            guests = db.query(Guest).filter(Guest.event_id == eid).filter(
                (Guest.name.ilike(f'%{query}%')) | (Guest.phone.ilike(f'%{query}%'))
            ).order_by(Guest.visual_id).all()
        else:
            guests = db.query(Guest).filter_by(event_id=eid).order_by(Guest.visual_id).all()

        return jsonify([{
            "id": g.id,
            "visual_id": g.visual_id,
            "name": g.name,
            "phone": g.phone,
            "qr_code_url": g.qr_code_url,
            "has_entered": g.has_entered,
            "checked_in_count": g.checked_in_count,
            "group_size": g.group_size,
            "entry_time": g.entry_time.strftime('%Y-%m-%d %H:%M:%S') if g.entry_time else 'N/A',
            "card_type": g.card_type
        } for g in guests])

# -------------------- download_excel --------------------

@app.route('/download_excel')
@admin_required
def download_excel():
    with get_db_session() as db:
        ev     = get_active_event(db)
        eid    = ev.id if ev else None
        guests = db.query(Guest).filter_by(event_id=eid).all()

    def ct(g): return (g.card_type or "").strip().lower()
    total_guests         = len(guests)
    single_cards         = sum(1 for g in guests if ct(g) == "single")
    double_cards         = sum(1 for g in guests if ct(g) == "double")
    family_cards         = sum(1 for g in guests if ct(g) == "family")
    total_family_allowed = sum(g.group_size for g in guests if ct(g) == "family")
    entered_guests       = sum(1 for g in guests if bool(g.has_entered))

    wb = Workbook()
    ws = wb.active
    ws.title      = "Guest Report"
    ws["A1"]      = "Guest Summary Report"
    ws["A1"].font = Font(size=14, bold=True)

    for row_num, (label, value) in enumerate([
        ("Total Guests", total_guests), ("Single Cards", single_cards),
        ("Double Cards", double_cards), ("Family Cards", family_cards),
        ("Total Allowed by Family Cards", total_family_allowed),
        ("Guests Entered", entered_guests),
        ("Guests Not Entered", total_guests - entered_guests),
    ], start=3):
        ws[f"A{row_num}"] = label; ws[f"B{row_num}"] = value
        ws[f"A{row_num}"].font = Font(bold=True)

    table_start = 11
    headers = ["ID", "Name", "Phone", "QR Code ID", "Has Entered", "Entry Time",
               "Visual ID", "Card Type", "Group Size", "WhatsApp", "RSVP", "AT SMS Sent"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=table_start, column=col, value=h).font = Font(bold=True)

    for i, g in enumerate(guests, start=table_start + 1):
        ws.cell(i,1,g.id);       ws.cell(i,2,g.name);      ws.cell(i,3,g.phone)
        ws.cell(i,4,g.qr_code_id)
        ws.cell(i,5,"Entered" if g.has_entered else "Not Entered")
        ws.cell(i,6,g.entry_time.strftime('%Y-%m-%d %H:%M:%S') if g.entry_time else "")
        ws.cell(i,7,g.visual_id); ws.cell(i,8,g.card_type); ws.cell(i,9,g.group_size)
        ws.cell(i,10,"Yes" if g.has_whatsapp else ("No" if g.has_whatsapp is False else "Unknown"))
        ws.cell(i,11,g.rsvp_status or "—")
        ws.cell(i,12,"Yes" if g.at_sms_sent else "No")

    fdr = table_start + 1; ldr = table_start + len(guests)
    if ldr >= fdr:
        rng = f"E{fdr}:E{ldr}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"Entered"'],
            fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"Not Entered"'],
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))

    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = (
            max((len(str(c.value)) for c in column if c.value), default=0) + 2)

    output = BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True, download_name="guest_report.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# -------------------- export_guests_simple --------------------

@app.route('/export_guests_simple')
@login_required
def export_guests_simple():
    """Export a simple guest list (Card No, Name, Phone, Card Type) as CSV or PDF."""
    fmt = request.args.get('format', 'pdf').lower()
    with get_db_session() as db:
        ev     = get_active_event(db)
        eid    = ev.id if ev else None
        guests = db.query(Guest).filter_by(event_id=eid).order_by(Guest.visual_id).all()
        rows = [
            (f"{g.visual_id:04d}", g.name or '', g.phone or '', (g.card_type or '').title())
            for g in guests
        ]

    headers = ['Card Number', 'Name', 'Phone Number', 'Card Type']

    if fmt == 'csv':
        si = StringIO()
        writer = csv.writer(si)
        writer.writerow(headers)
        writer.writerows(rows)
        output = make_response(si.getvalue())
        output.headers['Content-Disposition'] = 'attachment; filename=guests_list.csv'
        output.headers['Content-Type'] = 'text/csv'
        return output

    # Default: PDF
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    C_GREEN = colors.HexColor("#185a3f")
    C_GOLD  = colors.HexColor("#c9a84c")
    C_LIGHT = colors.HexColor("#f0f7f4")

    S_TITLE = ParagraphStyle('title', fontSize=16, textColor=C_GREEN,
                              alignment=TA_CENTER, spaceAfter=4)
    S_SUB   = ParagraphStyle('sub',   fontSize=9,  textColor=colors.grey,
                              alignment=TA_CENTER, spaceAfter=12)

    col_widths = [30*mm, 70*mm, 45*mm, 30*mm]
    table_data = [headers] + list(rows)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0),  C_GREEN),
        ('TEXTCOLOR',   (0,0), (-1,0),  colors.white),
        ('FONTNAME',    (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0),  10),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, C_LIGHT]),
        ('FONTSIZE',    (0,1), (-1,-1), 9),
        ('GRID',        (0,0), (-1,-1), 0.5, C_GOLD),
        ('ROWHEIGHT',   (0,0), (-1,-1), 7*mm),
    ]))

    story = [
        Paragraph('Guest List', S_TITLE),
        Paragraph(f'Generated {now_eat().strftime("%d %B %Y, %H:%M")} · {len(rows)} guests', S_SUB),
        t,
    ]
    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='guests_list.pdf',
                     mimetype='application/pdf')

# -------------------- zip_qr_codes_web --------------------

@app.route('/zip_qr_codes_web')
@login_required
def zip_qr_codes_web():
    with get_db_session() as db:
        ev     = get_active_event(db)
        eid    = ev.id if ev else None
        guests = db.query(Guest).filter_by(event_id=eid).all()
    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w') as zf:
        for guest in guests:
            if not guest.qr_code_url: continue
            try:
                zf.writestr(qr_filename_from_guest(guest),
                            download_from_supabase(QR_BUCKET, qr_filename_from_guest(guest)))
            except Exception as e:
                current_app.logger.warning(f"Could not fetch QR for {guest.name}: {e}")
    buf.seek(0)
    return send_file(buf, download_name='qr_codes.zip', as_attachment=True,
                     mimetype='application/zip')

# -------------------- edit_guest --------------------

@app.route('/edit_guest/<int:guest_id>', methods=['GET', 'POST'])
@admin_required
def edit_guest(guest_id):
    with get_db_session() as db:
        try:
            guest = db.query(Guest).filter(Guest.id == guest_id).first()
            if not guest:
                flash("Guest not found.", "danger")
                return redirect(url_for('view_all'))

            if request.method == 'POST':
                guest.name = request.form.get('name', guest.name).strip()
                guest.phone = to_whatsapp_number(request.form.get('phone', guest.phone))
                guest.has_entered = 'has_entered' in request.form
                new_card_type_raw = request.form.get('card_type', guest.card_type)
                group_size_raw = request.form.get('group_size', '').strip()
                new_card_type, _ = normalize_card_type(new_card_type_raw, group_size_raw or None)

                if new_card_type == "family":
                    try:
                        new_group_size = max(1, int(request.form.get('group_size', '').strip()))
                    except:
                        flash("Invalid group size for family card.", "danger")
                        return redirect(request.url)
                    if new_group_size < guest.checked_in_count:
                        flash(f"Group size cannot be less than checked-in count ({guest.checked_in_count}).", "danger")
                        return redirect(request.url)
                elif new_card_type == "single":
                    new_group_size = 1
                    if guest.checked_in_count > 1:
                        flash("Guest already used more scans than a single card allows.", "danger")
                        return redirect(request.url)
                elif new_card_type == "double":
                    new_group_size = 2
                    if guest.checked_in_count > 2:
                        flash("Guest already used more scans than a double card allows.", "danger")
                        return redirect(request.url)
                else:
                    flash("Unknown card type.", "danger")
                    return redirect(request.url)

                guest.card_type = new_card_type
                guest.group_size = new_group_size
                db.commit()
                flash('Guest updated successfully.', 'success')
                return redirect(url_for('view_all'))

            return render_template('edit_guest.html', guest=guest)

        except Exception as e:
            db.rollback()
            flash(f'Error updating guest: {e}', 'danger')
            current_app.logger.error(f"Error updating guest {guest_id}: {e}", exc_info=True)
            return redirect(url_for('view_all'))

@app.route('/scan_guests_data')
@login_required
def scan_guests_data():
    """Return all guests for the active event as JSON for offline scan caching."""
    with get_db_session() as db:
        ev  = get_active_event(db)
        eid = ev.id if ev else None
        guests = db.query(Guest).filter_by(event_id=eid).all()
        data = {}
        for g in guests:
            entry = {
                'id':               g.id,
                'name':             g.name or '',
                'visual_id':        g.visual_id or 0,
                'card_type':        g.card_type or 'single',
                'group_size':       g.group_size or 1,
                'checked_in_count': g.checked_in_count or 0,
                'has_entered':      bool(g.has_entered),
                'qr_code_id':       g.qr_code_id or '',
            }
            if g.qr_code_id:
                data[g.qr_code_id] = entry
            if g.visual_id:
                data[f'VID:{g.visual_id}'] = entry
    return jsonify(guests=data, event_name=ev.name if ev else '')


@app.route('/scan_qr')
@login_required
def scan_qr():
    return render_template('scan_qr.html')

# -------------------- delete_guest --------------------

@app.route('/delete_guest/<int:guest_id>', methods=['GET'])
@admin_required
def delete_guest(guest_id):
    with get_db_session() as db:
        try:
            guest = db.get(Guest, guest_id)
            if not guest:
                flash("Guest not found.", "danger")
                return redirect(url_for('view_all'))
            delete_from_supabase(QR_BUCKET,    qr_filename_from_guest(guest))
            delete_from_supabase(CARDS_BUCKET, card_filename_from_guest(guest))
            db.delete(guest)
            db.commit()
            flash('Guest and associated files deleted.', 'success')
        except Exception as e:
            db.rollback()
            flash(f'Error deleting guest: {e}', 'danger')
            current_app.logger.error(f"Error deleting guest {guest_id}: {e}", exc_info=True)
    return redirect(url_for('view_all'))

# -------------------- regenerate_qr_codes --------------------

@app.route('/regenerate_qr_codes')
@admin_required
def regenerate_qr_codes():
    with get_db_session() as db:
        try:
            ev     = get_active_event(db)
            eid    = ev.id if ev else None
            guests = db.query(Guest).filter_by(event_id=eid).all()
            for guest in guests:
                if guest.visual_id is None:
                    guest.visual_id = get_next_visual_id(db)
                qr_id    = f"GUEST-{guest.visual_id:04d}"
                qr_bytes = generate_qr_bytes(qr_id)
                qr_url   = upload_to_supabase(QR_BUCKET, qr_filename_from_guest(guest),
                                              qr_bytes, content_type="image/png")
                guest.qr_code_id  = qr_id
                guest.qr_code_url = qr_url
            db.commit()
            flash("QR codes regenerated.", "success")
        except Exception as e:
            db.rollback()
            flash(f"Error regenerating QR codes: {e}", "danger")
            current_app.logger.error(f"Error regenerating QR codes: {e}", exc_info=True)
    return redirect(url_for('view_all'))

# ===========================================================================
# CARD GENERATION — per-card API (avoids Gunicorn 30 s timeout)
# ===========================================================================

@app.route('/generate_guest_cards')
@admin_required
def generate_guest_cards():
    """Returns JSON list of visual_ids. Frontend calls /generate_card/<id> per guest."""
    if not os.path.exists(os.path.join("static", "Card Template.jpg")):
        return jsonify(success=False,
                       error="Card template not found at static/Card Template.jpg"), 500
    try:
        _bold_font_path()
    except FileNotFoundError as e:
        return jsonify(success=False, error=str(e)), 500

    with get_db_session() as db:
        ev     = get_active_event(db)
        eid    = ev.id if ev else None
        guests = db.query(Guest).filter_by(event_id=eid).order_by(Guest.visual_id).all()
        ids    = [g.visual_id for g in guests if g.qr_code_url]
    return jsonify(success=True, visual_ids=ids, total=len(ids))


@app.route('/generate_card/<int:visual_id>', methods=['POST'])
@login_required
def generate_card(visual_id):
    """Generate and upload card for ONE guest. Called by frontend in a loop."""
    with get_db_session() as db:
        guest = db.query(Guest).filter_by(visual_id=visual_id).first()
        if not guest:
            return jsonify(success=False, visual_id=visual_id,
                           error="Guest not found"), 404
        if not guest.qr_code_url:
            return jsonify(success=False, visual_id=visual_id,
                           error="No QR code — regenerate QR codes first"), 400
        ok = _render_and_upload_card(guest)
    if ok:
        return jsonify(success=True, visual_id=visual_id, name=guest.name)
    return jsonify(success=False, visual_id=visual_id,
                   error="Card rendering failed — check server logs"), 500


@app.route('/generate_cards_page')
@admin_required
def generate_cards_page():
    """Renders the progress-bar page that drives per-card generation."""
    return render_template('generate_cards.html')

# -------------------- download_card_by_id --------------------

@app.route('/download_card_by_id/<int:visual_id>')
@login_required
def download_card_by_id(visual_id):
    with get_db_session() as db:
        try:
            guest = db.query(Guest).filter_by(visual_id=visual_id).first()
            if not guest:
                flash("Guest not found.", "danger")
                return redirect(url_for('view_all'))
            if not os.path.exists(os.path.join("static", "Card Template.jpg")):
                flash("Card template missing.", "danger")
                return redirect(url_for('view_all'))

            try:
                qr_data = download_from_supabase(QR_BUCKET, qr_filename_from_guest(guest))
            except Exception:
                qr_data = generate_qr_bytes(guest.qr_code_id)

            qr_img = Image.open(BytesIO(qr_data))
            img    = _draw_card(guest, qr_img)
            buf    = BytesIO()
            img.save(buf, format="JPEG", quality=92)
            buf.seek(0)
            return send_file(buf, as_attachment=True,
                             download_name=f"Guest-{guest.visual_id:04d}.jpg",
                             mimetype="image/jpeg")
        except Exception as e:
            flash(f"Error generating card: {e}", "danger")
            current_app.logger.error(f"Error downloading card: {e}", exc_info=True)
            return redirect(url_for('view_all'))

# -------------------- download_all_cards --------------------

@app.route('/download_all_cards')
@login_required
def download_all_cards():
    """
    Regenerates every guest card on the fly and bundles them into a ZIP.
    Uses per-event card template if available.
    """
    try:
        if not os.path.exists(os.path.join("static", "Card Template.jpg")):
            flash("Card template missing — cannot generate cards.", "danger")
            return redirect(url_for('view_all'))

        with get_db_session() as db:
            ev     = get_active_event(db)
            eid    = ev.id if ev else None
            guests = db.query(Guest).filter_by(event_id=eid).order_by(Guest.visual_id).all()
            # Snapshot all data we need while session is open
            guest_snapshots = [
                {
                    "visual_id":   g.visual_id,
                    "name":        g.name,
                    "phone":       g.phone,
                    "card_type":   g.card_type,
                    "group_size":  g.group_size,
                    "qr_code_id":  g.qr_code_id,
                    "qr_filename": qr_filename_from_guest(g),
                    "card_fname":  card_filename_from_guest(g),
                }
                for g in guests
            ]
            # Fetch event template bytes once (outside the per-guest loop)
            event_template_bytes = _get_event_template_bytes(ev)

        if not guest_snapshots:
            flash("No guests found.", "warning")
            return redirect(url_for('view_all'))

        zip_buffer = BytesIO()
        count  = 0
        errors = []

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for snap in guest_snapshots:
                try:
                    try:
                        qr_data = download_from_supabase(QR_BUCKET, snap["qr_filename"])
                    except Exception:
                        qr_data = generate_qr_bytes(snap["qr_code_id"])

                    # Build a lightweight guest-like object for _draw_card
                    class _G:
                        pass
                    g = _G()
                    g.visual_id  = snap["visual_id"]
                    g.name       = snap["name"]
                    g.phone      = snap["phone"]
                    g.card_type  = snap["card_type"]
                    g.group_size = snap["group_size"]

                    qr_img = Image.open(BytesIO(qr_data))
                    card   = _draw_card(g, qr_img, template_bytes=event_template_bytes)
                    buf    = BytesIO()
                    card.save(buf, format="JPEG", quality=92)
                    zf.writestr(snap["card_fname"], buf.getvalue())
                    count += 1

                except Exception as e:
                    current_app.logger.warning(
                        f"Skipped card for guest {snap['visual_id']} ({snap['name']}): {e}"
                    )
                    errors.append(snap["name"])

        if count == 0:
            flash("Could not generate any cards. Check the card template and logs.", "danger")
            return redirect(url_for('view_all'))

        if errors:
            flash(f"Downloaded {count} cards. Skipped {len(errors)}: {', '.join(errors)}", "warning")

        zip_buffer.seek(0)
        response = make_response(zip_buffer.read())
        response.headers['Content-Type'] = 'application/zip'
        response.headers['Content-Disposition'] = 'attachment; filename="invitation_cards.zip"'
        return response

    except Exception as e:
        current_app.logger.exception(f"download_all_cards failed: {e}")
        flash(f"Error: {e}", "danger")
        return redirect(url_for('view_all'))

# -------------------- guest_report --------------------

@app.route('/guest_report_data')
@login_required
def guest_report_data():
    with get_db_session() as db:
        ev     = get_active_event(db)
        eid    = ev.id if ev else None
        guests = db.query(Guest).filter_by(event_id=eid).order_by(Guest.visual_id).all()

    def g_dict(g):
        return {
            "name":             g.name,
            "visual_id":        g.visual_id,
            "card_type":        g.card_type,
            "group_size":       g.group_size,
            "checked_in_count": g.checked_in_count,
            "rsvp_status":      g.rsvp_status,
            "entry_time":       g.entry_time.strftime('%H:%M') if g.entry_time else None,
        }

    total       = len(guests)
    entered     = [g for g in guests if (g.checked_in_count or 0) >= 1]
    not_entered = [g for g in guests if (g.checked_in_count or 0) == 0]
    attending   = [g for g in guests if g.rsvp_status == 'attending']
    declined    = [g for g in guests if g.rsvp_status == 'not_attending']
    no_rsvp     = [g for g in guests if not g.rsvp_status]
    wa_sent     = [g for g in guests if g.whatsapp_sent]
    wa_pending  = [g for g in guests if not g.whatsapp_sent]
    sms_sent    = [g for g in guests if g.at_sms_sent]

    return jsonify({
        "total_guests":       total,
        "single_cards":       sum(1 for g in guests if (g.card_type or '') == 'single'),
        "double_cards":       sum(1 for g in guests if (g.card_type or '') == 'double'),
        "family_cards":       sum(1 for g in guests if (g.card_type or '') == 'family'),
        "entered_guests":     len(entered),
        "not_entered_guests": len(not_entered),
        "attending":          len(attending),
        "not_attending":      len(declined),
        "no_rsvp":            len(no_rsvp),
        "wa_sent":            len(wa_sent),
        "wa_pending":         len(wa_pending),
        "sms_sent":           len(sms_sent),
        "entered_list":       [g_dict(g) for g in entered],
        "not_entered_list":   [g_dict(g) for g in not_entered],
        "attending_list":     [g_dict(g) for g in attending],
        "declined_list":      [g_dict(g) for g in declined],
        "no_rsvp_list":       [g_dict(g) for g in no_rsvp],
        "wa_sent_list":       [g_dict(g) for g in wa_sent],
        "wa_pending_list":    [g_dict(g) for g in wa_pending],
    })


@app.route('/guest_report')
@login_required
def guest_report():
    return render_template('guest_report.html')

# -------------------- clear_all_data --------------------

@app.route('/clear_all_data', methods=['GET'])
@admin_required
def clear_all_data():
    with get_db_session() as db:
        try:
            ev  = get_active_event(db)
            eid = ev.id if ev else None
            guests = db.query(Guest).filter_by(event_id=eid).all()
            for guest in guests:
                delete_from_supabase(QR_BUCKET,    qr_filename_from_guest(guest))
                delete_from_supabase(CARDS_BUCKET, card_filename_from_guest(guest))
            num_deleted = db.query(Guest).filter_by(event_id=eid).delete()
            db.commit()
            flash(f"Successfully deleted {num_deleted} guests from this event.", "success")
        except Exception as e:
            db.rollback()
            flash(f"An error occurred while clearing data: {e}", "danger")
            current_app.logger.error(f"Error clearing all data: {e}", exc_info=True)
    return redirect(url_for('view_all'))

# ===========================================================================
# WEBHOOK  (RSVP receiver)
# ===========================================================================

@app.route('/webhook/whatsapp', methods=['GET', 'POST'])
def whatsapp_webhook():
    if request.method == 'GET':
        mode      = request.args.get('hub.mode')
        token     = request.args.get('hub.verify_token')
        challenge = request.args.get('hub.challenge')
        if mode == 'subscribe' and token == WHATSAPP_VERIFY_TOKEN:
            current_app.logger.info("Webhook verified by Meta.")
            return challenge, 200
        return "Forbidden", 403

    try:
        data = request.get_json()
        current_app.logger.info(f"Webhook payload: {data}")
        for entry in data.get('entry', []):
            for change in entry.get('changes', []):
                for msg in change.get('value', {}).get('messages', []):
                    from_number = msg.get('from')
                    if msg.get('type') == 'button':
                        _handle_rsvp(from_number,
                                     msg.get('button', {}).get('text', '').strip())
                    elif msg.get('type') == 'interactive':
                        inter = msg.get('interactive', {})
                        if inter.get('type') == 'button_reply':
                            _handle_rsvp(from_number,
                                         inter.get('button_reply', {}).get('title', '').strip())
    except Exception as e:
        current_app.logger.error(f"Webhook error: {e}", exc_info=True)
    return "OK", 200


def _handle_rsvp(from_number: str, button_text: str):
    text = button_text.lower()
    if any(x in text for x in ['nitakuwepo', "i'll be there", 'attending']):
        rsvp_status = 'attending'
    elif any(x in text for x in ['sitakuwepo', "can't make it", 'not attending']):
        rsvp_status = 'not_attending'
    else:
        current_app.logger.warning(f"Unknown button text from {from_number}: {button_text}")
        return

    raw = str(from_number).strip().lstrip('+')
    variants = {raw, f"+{raw}"}
    if raw.startswith("255") and len(raw) >= 11:
        local9 = raw[3:]
        variants.update({local9, f"0{local9}", f"+255{local9}"})
    if raw.startswith("0") and len(raw) == 10:
        local9 = raw[1:]
        variants.update({f"255{local9}", f"+255{local9}", local9})

    with get_db_session() as db:
        guest = db.query(Guest).filter(Guest.phone.in_(variants)).first()
        if not guest:
            current_app.logger.warning(
                f"No guest found for: {from_number} (tried: {variants})")
            return
        guest.rsvp_status = rsvp_status
        guest.rsvp_at     = now_eat()
        db.commit()
        current_app.logger.info(
            f"RSVP saved: {guest.name} → {rsvp_status}")

# ===========================================================================
# UNIFIED SEND ENGINE
# ===========================================================================

def _send_to_guest(guest, db, send_wa=True, send_sms=True, event=None):
    """
    Deliver invitation to one guest.

    Parameters
    ----------
    send_wa  : bool  — attempt WhatsApp delivery
    send_sms : bool  — attempt Africa's Talking SMS delivery
    event    : Event — active event; used for per-event WA template + SMS copy

    Returns
    -------
    dict with keys: wa, sms, overall, message
      wa / sms values: "sent" | "skipped" | "failed" | "invalid" | "not_configured"
      overall:         "success" | "partial" | "failed"
    """
    now        = now_eat()
    wa_status  = "skipped"
    sms_status = "skipped"
    messages   = []

    phone = to_whatsapp_number(guest.phone)
    if not phone:
        return {"wa": "failed", "sms": "failed",
                "overall": "failed", "message": "No valid phone number."}

    # ── WhatsApp ──────────────────────────────────────────────────────────
    if send_wa:
        print(f"[WA DEBUG] Starting WA send for guest: {guest.name} | phone: {phone}", flush=True)
        try:
            card_fname = card_filename_from_guest(guest)
            print(f"[WA DEBUG] Card filename: {card_fname}", flush=True)

            # Try fetching from Supabase first, fall back to regenerating
            card_bytes = None
            try:
                card_bytes = download_from_supabase(CARDS_BUCKET, card_fname)
                print(f"[WA DEBUG] Card fetched from Supabase: {len(card_bytes)} bytes", flush=True)
            except Exception as supa_err:
                print(f"[WA DEBUG] Supabase fetch failed ({supa_err}), regenerating card...", flush=True)
                card_bytes = _generate_card_bytes(guest, event=event)
                if card_bytes:
                    print(f"[WA DEBUG] Card regenerated: {len(card_bytes)} bytes", flush=True)
                    try:
                        upload_to_supabase(CARDS_BUCKET, card_fname, card_bytes,
                                           content_type="image/jpeg")
                    except Exception as up_err:
                        print(f"[WA DEBUG] Re-upload to Supabase failed (non-fatal): {up_err}", flush=True)
                else:
                    print(f"[WA DEBUG] Card regeneration also failed!", flush=True)

            if not card_bytes:
                raise ValueError("Could not retrieve or generate card image.")

            # ── Resolve per-event WhatsApp credentials ────────────────────
            # Each field falls back to the global env var if the event has no override.
            wa_template_name     = (event.wa_template_name     if event and event.wa_template_name
                                    else None)
            wa_template_language = (event.wa_template_language if event and event.wa_template_language
                                    else None)
            wa_phone_number_id   = (event.wa_phone_number_id   if event and event.wa_phone_number_id
                                    else None)
            wa_access_token      = (event.wa_access_token      if event and event.wa_access_token
                                    else None)

            print(
                f"[WA DEBUG] Using template='{wa_template_name or 'DEFAULT'}' "
                f"lang='{wa_template_language or 'DEFAULT'}' "
                f"phone_id='{wa_phone_number_id or 'ENV'}' "
                f"token={'CUSTOM' if wa_access_token else 'ENV'}",
                flush=True,
            )
            print(f"[WA DEBUG] Calling send_guest_card...", flush=True)

            wa_result = send_guest_card(
                to=phone,
                guest_name=guest.name or "Guest",
                visual_id=guest.visual_id,
                card_type=guest.card_type,
                image_bytes=card_bytes,
                filename=card_fname,
                # ── per-event overrides (None = use whatsapp.py defaults) ──
                template_name=wa_template_name,
                template_language=wa_template_language,
                phone_number_id=wa_phone_number_id,
                access_token=wa_access_token,
            )
            print(f"[WA DEBUG] send_guest_card result: {wa_result}", flush=True)

            if wa_result.get("status") == "invalid_number":
                guest.has_whatsapp        = False
                guest.whatsapp_checked_at = now
                guest.whatsapp_sent       = False
                guest.whatsapp_error      = "Not on WhatsApp"
                wa_status = "invalid"
                messages.append("WhatsApp: not on platform.")
            else:
                guest.whatsapp_sent    = True
                guest.whatsapp_sent_at = now
                guest.whatsapp_error   = None
                wa_status = "sent"
                messages.append("WhatsApp: sent.")
        except Exception as e:
            import traceback
            err_str = str(e)[:500]
            tb      = traceback.format_exc()
            print(f"[WA ERROR] WA send failed for {guest.name}: {err_str}", flush=True)
            print(f"[WA ERROR] Traceback:\n{tb}", flush=True)
            guest.whatsapp_sent  = False
            guest.whatsapp_error = err_str
            wa_status = "failed"
            messages.append(f"WhatsApp failed: {err_str}")
            logging.error(f"WA send failed for {guest.name}: {e}", exc_info=True)
    # wa_status stays "skipped" if send_wa is False

    # ── Africa's Talking SMS ──────────────────────────────────────────────
    if send_sms:
        if not at_configured():
            sms_status = "not_configured"
            messages.append("SMS: Africa's Talking not configured.")
        else:
            try:
                sms_result = at_send_sms(phone, build_sms_message(guest, event=event))
                if sms_result.get("success"):
                    guest.at_sms_sent    = True
                    guest.at_sms_error   = None
                    guest.at_sms_sent_at = now
                    sms_status = "sent"
                    messages.append("SMS: sent.")
                else:
                    err_str = sms_result.get("error", "Unknown SMS error")[:500]
                    guest.at_sms_sent  = False
                    guest.at_sms_error = err_str
                    sms_status = "failed"
                    messages.append(f"SMS failed: {err_str}")
            except Exception as e:
                err_str = str(e)[:500]
                guest.at_sms_sent  = False
                guest.at_sms_error = err_str
                sms_status = "failed"
                messages.append(f"SMS error: {err_str}")
                logging.error(f"SMS send failed for {guest.name}: {e}", exc_info=True)
    # sms_status stays "skipped" if send_sms is False

    db.commit()

    # Overall outcome — only consider channels that were actually requested
    active = []
    if send_wa:  active.append(wa_status)
    if send_sms: active.append(sms_status)

    if any(s == "sent" for s in active):
        overall = "success"
    elif all(s in ("failed", "invalid", "not_configured") for s in active):
        overall = "failed"
    else:
        overall = "partial"

    return {"wa": wa_status, "sms": sms_status,
            "overall": overall, "message": " | ".join(messages)}


# ── Unified single — WA + SMS ─────────────────────────────────────────────────

@app.route('/send_unified_single/<int:guest_id>', methods=['POST'])
@login_required
def send_unified_single(guest_id):
    with get_db_session() as db:
        guest = db.get(Guest, guest_id)
        if not guest:
            return jsonify(success=False, message="Guest not found.")
        ev     = get_active_event(db)
        result = _send_to_guest(guest, db, send_wa=True, send_sms=True, event=ev)
        return jsonify(success=(result["overall"] != "failed"),
                       overall=result["overall"], wa=result["wa"],
                       sms=result["sms"], message=result["message"],
                       guest_id=guest_id)


# ── Unified bulk — WA + SMS ───────────────────────────────────────────────────

@app.route('/send_unified_bulk', methods=['POST'])
@login_required
def send_unified_bulk():
    data   = request.get_json() or {}
    resend = data.get('resend', False)
    with get_db_session() as db:
        if resend:
            ev  = get_active_event(db)
            eid = ev.id if ev else None
            guests = db.query(Guest).filter_by(event_id=eid).order_by(Guest.visual_id).all()
        else:
            ev  = get_active_event(db)
            eid = ev.id if ev else None
            guests = db.query(Guest).filter(
                Guest.event_id == eid,
                ((Guest.whatsapp_sent == False) | (Guest.whatsapp_sent == None)),
                ((Guest.at_sms_sent   == False) | (Guest.at_sms_sent   == None))
            ).order_by(Guest.visual_id).all()

        totals = {"total": len(guests), "wa_sent": 0, "wa_failed": 0,
                  "sms_sent": 0, "sms_failed": 0, "errors": []}
        for guest in guests:
            result = _send_to_guest(guest, db, send_wa=True, send_sms=True, event=ev)
            if result["wa"]  == "sent":                  totals["wa_sent"]    += 1
            elif result["wa"]  in ("failed", "invalid"): totals["wa_failed"]  += 1
            if result["sms"] == "sent":                  totals["sms_sent"]   += 1
            elif result["sms"] == "failed":              totals["sms_failed"] += 1
            if result["overall"] == "failed":
                totals["errors"].append({"name": guest.name, "error": result["message"]})
            time.sleep(0.1)
        return jsonify(totals)


# ── WhatsApp only — single ────────────────────────────────────────────────────

@app.route('/send_card_single/<int:guest_id>', methods=['POST'])
@login_required
def send_card_single(guest_id):
    with get_db_session() as db:
        guest = db.get(Guest, guest_id)
        if not guest:
            return jsonify(success=False, message="Guest not found.")
        ev     = get_active_event(db)
        result = _send_to_guest(guest, db, send_wa=True, send_sms=False, event=ev)
        return jsonify(success=(result["wa"] == "sent"),
                       message=result["message"], guest_id=guest_id)


# ── WhatsApp only — bulk ──────────────────────────────────────────────────────

@app.route('/send_cards_bulk', methods=['POST'])
@login_required
def send_cards_bulk():
    data   = request.get_json() or {}
    resend = data.get('resend', False)
    with get_db_session() as db:
        if resend:
            ev  = get_active_event(db)
            eid = ev.id if ev else None
            guests = db.query(Guest).filter_by(event_id=eid).order_by(Guest.visual_id).all()
        else:
            ev  = get_active_event(db)
            eid = ev.id if ev else None
            guests = db.query(Guest).filter(
                Guest.event_id == eid,
                (Guest.whatsapp_sent == False) | (Guest.whatsapp_sent == None)
            ).order_by(Guest.visual_id).all()

        sent = failed = 0
        errors = []
        for guest in guests:
            result = _send_to_guest(guest, db, send_wa=True, send_sms=False, event=ev)
            if result["wa"] == "sent":
                sent += 1
            else:
                failed += 1
                if result["overall"] == "failed":
                    errors.append({"name": guest.name, "error": result["message"]})
            time.sleep(0.1)
        return jsonify(total=len(guests), sent=sent, failed=failed, errors=errors)


# ── SMS only — single ─────────────────────────────────────────────────────────

@app.route('/send_at_sms_single/<int:guest_id>', methods=['POST'])
@login_required
def send_at_sms_single(guest_id):
    with get_db_session() as db:
        guest = db.get(Guest, guest_id)
        if not guest:
            return jsonify(success=False, message="Guest not found."), 404
        ev     = get_active_event(db)
        result = _send_to_guest(guest, db, send_wa=False, send_sms=True, event=ev)
        return jsonify(success=(result["sms"] == "sent"),
                       message=result["message"])


# ── SMS only — bulk ───────────────────────────────────────────────────────────

@app.route('/send_at_sms_bulk', methods=['POST'])
@login_required
def send_at_sms_bulk():
    data   = request.get_json() or {}
    resend = data.get("resend", False)
    with get_db_session() as db:
        if resend:
            ev  = get_active_event(db)
            eid = ev.id if ev else None
            guests = db.query(Guest).filter_by(event_id=eid).all()
        else:
            ev  = get_active_event(db)
            eid = ev.id if ev else None
            guests = db.query(Guest).filter(
                Guest.event_id == eid,
                (Guest.at_sms_sent == None) | (Guest.at_sms_sent == False)
            ).all()

        sent_count = failed_count = 0
        errors = []
        for guest in guests:
            result = _send_to_guest(guest, db, send_wa=False, send_sms=True, event=ev)
            if result["sms"] == "sent":
                sent_count += 1
            else:
                failed_count += 1
                if result["overall"] == "failed":
                    errors.append({"name": guest.name, "error": result["message"]})
            time.sleep(0.1)
        return jsonify(total=len(guests), sent=sent_count,
                       failed=failed_count, errors=errors)


# -------------------- send_cards page --------------------

@app.route('/send_cards', methods=['GET'])
@login_required
def send_cards():
    with get_db_session() as db:
        ev     = get_active_event(db)
        eid    = ev.id if ev else None
        guests = db.query(Guest).filter_by(event_id=eid).order_by(Guest.visual_id).all()
        total         = len(guests)
        sent          = sum(1 for g in guests if g.whatsapp_sent)
        failed        = sum(1 for g in guests if g.whatsapp_error and not g.whatsapp_sent)
        pending       = total - sent
        attending     = sum(1 for g in guests if g.rsvp_status == 'attending')
        not_attending = sum(1 for g in guests if g.rsvp_status == 'not_attending')
        no_rsvp       = total - attending - not_attending
        at_sms_sent_count   = sum(1 for g in guests if g.at_sms_sent)
        at_sms_failed_count = sum(1 for g in guests if g.at_sms_error and not g.at_sms_sent)
        wa_checked          = sum(1 for g in guests if g.has_whatsapp is not None)
        no_whatsapp_count   = sum(1 for g in guests if g.has_whatsapp is False)

        return render_template(
            'send_cards.html',
            guests=guests,
            total=total, sent=sent, failed=failed, pending=pending,
            attending=attending, not_attending=not_attending, no_rsvp=no_rsvp,
            wa_checked=wa_checked, no_whatsapp_count=no_whatsapp_count,
            at_configured=at_configured(),
            at_sms_sent_count=at_sms_sent_count,
            at_sms_failed_count=at_sms_failed_count,
            sms_enabled=at_configured(),
            sms_sent_count=at_sms_sent_count,
        )

# ===========================================================================
# Misc / public routes
# ===========================================================================

@app.route("/privacy")
def privacy():
    return render_template("privacy.html")


@app.route("/data-deletion")
def data_deletion():
    return """
    <html>
    <head><title>Data Deletion - SwiftInvite</title></head>
    <body style="font-family: Arial; margin: 40px;">
    <h1>User Data Deletion</h1>
    <p>If you would like to delete your data from SwiftInvite,
       please follow the instructions below:</p>
    <ol>
        <li>Send an email to: <strong>swiftinvite25@gmail.com</strong></li>
        <li>Include your phone number or identifier used in the app</li>
        <li>We will process your request within 7 days</li>
    </ol>
    <p>Alternatively, you may contact us directly for assistance.</p>
    </body>
    </html>
    """

@app.route('/download_client_report')
@login_required
def download_client_report():
    """
    Client-facing PDF: stat summary + two clean tables
    (checked-in guests and not-checked-in guests).
    NO RSVP data anywhere in this document.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    with get_db_session() as db:
        ev     = get_active_event(db)
        eid    = ev.id if ev else None
        guests = db.query(Guest).filter_by(event_id=eid).order_by(Guest.visual_id).all()

    total         = len(guests)
    entered       = [g for g in guests if (g.checked_in_count or 0) >= 1]
    not_entered   = [g for g in guests if (g.checked_in_count or 0) == 0]
    single_count  = sum(1 for g in guests if (g.card_type or '') == 'single')
    double_count  = sum(1 for g in guests if (g.card_type or '') == 'double')
    family_count  = sum(1 for g in guests if (g.card_type or '') == 'family')
    total_allowed = sum(g.group_size or 1 for g in guests)
    generated_at  = now_eat().strftime("%d %B %Y, %H:%M")

    C_GREEN  = colors.HexColor("#185a3f")
    C_GOLD   = colors.HexColor("#c9a84c")
    C_RED    = colors.HexColor("#dc2626")
    C_LIGHT  = colors.HexColor("#f0f7f4")
    C_BORDER = colors.HexColor("#d1d5db")
    C_GREY   = colors.HexColor("#6b7280")
    C_INK    = colors.HexColor("#1f2937")
    C_WHITE  = colors.white
    C_ROW2   = colors.HexColor("#fafafa")

    W = A4[0] - 40 * mm

    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    S_TITLE  = ps('tt', fontName='Helvetica-Bold', fontSize=22, textColor=C_GREEN,  alignment=TA_CENTER, spaceAfter=3)
    S_META   = ps('me', fontName='Helvetica',      fontSize=8,  textColor=C_GREY,   alignment=TA_CENTER, spaceAfter=12)
    S_SEC    = ps('se', fontName='Helvetica-Bold', fontSize=12, textColor=C_GREEN,  spaceBefore=14, spaceAfter=5)
    S_CELL   = ps('ce', fontName='Helvetica',      fontSize=8.5, textColor=C_INK,   leading=12)
    S_CELL_B = ps('cb', fontName='Helvetica-Bold', fontSize=8.5, textColor=C_INK,   leading=12)
    S_HEAD   = ps('hd', fontName='Helvetica-Bold', fontSize=8,  textColor=C_WHITE,  alignment=TA_CENTER)

    def stat_table(items):
        cw = W / len(items)
        data = [[
            Paragraph(
                f"<font size='20' color='{c.hexval()}'><b>{v}</b></font><br/>"
                f"<font size='7' color='#6b7280'>{lbl}</font>",
                ps(f'sr{i}', fontName='Helvetica-Bold', fontSize=20,
                   textColor=c, alignment=TA_CENTER, leading=24))
            for i, (lbl, v, c) in enumerate(items)
        ]]
        tbl = Table(data, colWidths=[cw] * len(items))
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,-1), C_LIGHT),
            ('BOX',           (0,0), (-1,-1), 0.5, C_BORDER),
            ('INNERGRID',     (0,0), (-1,-1), 0.3, C_BORDER),
            ('TOPPADDING',    (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ]))
        return tbl

    def guest_table(guest_list, cols, col_widths, row_fn, empty_msg="None"):
        header = [Paragraph(c, S_HEAD) for c in cols]
        rows   = [header]
        if guest_list:
            for i, g in enumerate(guest_list):
                rows.append(row_fn(g, i))
        else:
            rows.append([Paragraph(f"<i>{empty_msg}</i>", S_CELL)]
                        + [Paragraph('', S_CELL)] * (len(cols) - 1))
        tbl = Table(rows, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),  (-1,0),  C_GREEN),
            ('ROWBACKGROUNDS',(0,1),  (-1,-1), [C_WHITE, C_ROW2]),
            ('GRID',          (0,0),  (-1,-1), 0.3, C_BORDER),
            ('LEFTPADDING',   (0,0),  (-1,-1), 6),
            ('RIGHTPADDING',  (0,0),  (-1,-1), 6),
            ('TOPPADDING',    (0,0),  (-1,-1), 5),
            ('BOTTOMPADDING', (0,0),  (-1,-1), 5),
            ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
            ('FONTNAME',      (0,1),  (-1,-1), 'Helvetica'),
            ('FONTSIZE',      (0,1),  (-1,-1), 8.5),
        ]))
        return tbl

    def in_row(g, i):
        et = g.entry_time.strftime('%H:%M') if g.entry_time else '—'
        return [
            Paragraph(f"{g.visual_id:04d}", S_CELL_B),
            Paragraph(f"<b>{g.name or '—'}</b>", S_CELL),
            Paragraph((g.card_type or 'single').title(), S_CELL),
            Paragraph(f"{g.checked_in_count or 0}/{g.group_size or 1}", S_CELL),
            Paragraph(et, S_CELL),
        ]

    def nin_row(g, i):
        return [
            Paragraph(f"{g.visual_id:04d}", S_CELL_B),
            Paragraph(f"<b>{g.name or '—'}</b>", S_CELL),
            Paragraph((g.card_type or 'single').title(), S_CELL),
            Paragraph(str(g.group_size or 1), S_CELL),
        ]

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=16*mm,
        title="Guest Report", author="SwiftInvite",
    )

    def on_page(canvas, doc):
        canvas.saveState()
        w, h = A4
        canvas.setFillColor(C_GREEN)
        canvas.rect(0, h - 11*mm, w, 11*mm, fill=1, stroke=0)
        canvas.setFillColor(C_WHITE)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawCentredString(w/2, h - 7.5*mm, "GUEST REPORT — CONFIDENTIAL")
        canvas.setFillColor(colors.HexColor("#f3f4f6"))
        canvas.rect(0, 0, w, 8*mm, fill=1, stroke=0)
        canvas.setFillColor(C_GREY)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(20*mm, 3*mm, f"Generated by SwiftInvite · {generated_at}")
        canvas.drawRightString(w - 20*mm, 3*mm, f"Page {doc.page}")
        canvas.restoreState()

    story = []
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph("Guest Report", S_TITLE))
    story.append(Paragraph(f"Prepared on {generated_at}", S_META))
    story.append(HRFlowable(width="100%", thickness=1.5, color=C_GOLD, spaceAfter=10))

    story.append(Paragraph("Summary", S_SEC))
    story.append(stat_table([
        ("Total Guests",  total,           C_GREEN),
        ("Total Allowed", total_allowed,   C_GREEN),
        ("Checked In",    len(entered),    C_GREEN),
        ("Not In Yet",    len(not_entered), C_RED),
    ]))
    story.append(Spacer(1, 5))
    story.append(stat_table([
        ("Single Cards", single_count, C_GOLD),
        ("Double Cards", double_count, C_GOLD),
        ("Family Cards", family_count, C_GOLD),
    ]))

    story.append(Paragraph(f"Checked In — {len(entered)} of {total} guests", S_SEC))
    story.append(HRFlowable(width="100%", thickness=0.4, color=C_BORDER, spaceAfter=5))
    story.append(guest_table(
        entered,
        ["#", "Guest Name", "Card Type", "Entries", "Time In"],
        [14*mm, W - 14*mm - 22*mm - 18*mm - 22*mm, 22*mm, 18*mm, 22*mm],
        in_row,
        "No guests have checked in yet.",
    ))

    story.append(Paragraph(f"Not Yet Checked In — {len(not_entered)} guests", S_SEC))
    story.append(HRFlowable(width="100%", thickness=0.4, color=C_BORDER, spaceAfter=5))
    story.append(guest_table(
        not_entered,
        ["#", "Guest Name", "Card Type", "Allowed"],
        [14*mm, W - 14*mm - 22*mm - 18*mm, 22*mm, 18*mm],
        nin_row,
        "All guests have checked in!",
    ))

    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=0.4, color=C_BORDER))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        "This report was generated automatically by SwiftInvite. "
        "All guest data is confidential.",
        ps('ft', fontName='Helvetica', fontSize=7.5,
           textColor=C_GREY, alignment=TA_CENTER)))

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    buf.seek(0)

    filename = f"Guest_Report_{now_eat().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype='application/pdf')


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)